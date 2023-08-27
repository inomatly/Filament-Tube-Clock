import os
from bs4 import BeautifulSoup
import openpyxl


def extract_data_from_html(file_path):
    with open(file_path, 'r', encoding="utf-8") as file:
        soup = BeautifulSoup(file, "html.parser")
        
        modules = []
        coverages = []

        # tbody内のすべての行（tr）を取得
        rows = soup.find("tbody").findAll("tr", class_="file")

        for row in rows:
            module_cell = row.find("td", class_="name")
            coverage_cell = row.find("td", class_="right")

            module_name = module_cell.a.string
            coverage_value = coverage_cell.get_text().strip('%')

            modules.append(module_name)
            coverages.append(coverage_value)
        
        # 合計のカバレッジを取得
        total_row = soup.find("tfoot").find("tr", class_="total")
        total_coverage_cell = total_row.find("td", class_="right")
        total_coverage_value = total_coverage_cell.get_text().strip('%')
        
        modules.append("Total")
        coverages.append(total_coverage_value)

    return modules, coverages


def write_to_excel(modules, coverages, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    for i, (module, coverage) in enumerate(zip(modules, coverages), 1):
        ws[f'A{i}'] = module
        ws[f'B{i}'] = int(coverage)
        ws[f'C{i}'] = "%"

    wb.save(output_path)


def main():
    # パスの設定
    current_dir = os.path.dirname(os.path.abspath(__file__))
    html_path = os.path.join(current_dir, "reports", "cov.html", "index.html")
    output_excel_path = os.path.join(current_dir, "./reports/coverage.xlsx")

    modules, coverages = extract_data_from_html(html_path)

    # 標準出力
    for module, coverage in zip(modules, coverages):
        print(f"{module}: {coverage}%")

    # エクセルに書き込む
    write_to_excel(modules, coverages, output_excel_path)


if __name__ == "__main__":
    main()
